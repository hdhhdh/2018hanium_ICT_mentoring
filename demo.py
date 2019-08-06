import os
import cv2
import dlib
import numpy as np
import argparse
import time
import threading
import random
import socket
import glob
import matplotlib.pyplot as plt
import sys
 
from openpyxl import Workbook
from openpyxl import load_workbook
from PyQt5.QtCore import QTimer
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import QApplication, QDialog, QWidget, QVBoxLayout, QHBoxLayout, QBoxLayout
from PyQt5.uic import loadUi
from PyQt5.uic.properties import QtCore
 
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotwidgetfile import matplotwidgetFile
 
from contextlib import contextmanager
from wide_resnet import WideResNet
from keras.utils.data_utils import get_file
 
global myimge
global Agecnt
global Range
global exitflag
global sock
global sock_on
global th_flag
global testflag
global name
global xl_row
 
myimge = cv2.imread("waiting.png")
Agecnt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
Agevideo = [r"C:\ProgramData\Age\age-gender-estimation-master\ad\teenF",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\YouthF",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\middleF",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\oldF",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\teenM",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\YouthM",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\middleM",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\oldM",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\079war01.mp4",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\068war02.mp4",
            r"C:\ProgramData\Age\age-gender-estimation-master\ad\109war03.mp4"]
 
Range = ['teenF', 'YouthF', 'middleF', 'oldF', 'teenM', 'YouthM', 'middleM', 'oldM', 'warning']
exitflag = 0
sock_on = 1
th_flag = 0
testflag = 0
 
if sock_on == 1:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR,1)
    sock.connect(('192.168.43.247', 5005))
 
 
class Video(object):
    def __init__(self, path):
        self.path = path
 
    def play(self):
        from os import startfile
        startfile(self.path)
 
 
class Movie_MP4(Video):
    type = "MP4"
 
 
class Life2Coding(QDialog):
 
    def __init__(self):
        super(Life2Coding, self).__init__()
        loadUi('life2coding.ui', self)
 
        self.setGeometry(900, 30, 1006, 729)
        # self.timer = QTimer(self)
        self.imgtimer = QTimer(self)
 
        self.startButton.clicked.connect(self.start_webcam)
        self.stopButton.clicked.connect(self.stop_webcam)
        self.warningButton.clicked.connect(self.warning_event)
        self.resetButton.clicked.connect(self.reset)
        self.fireButton.clicked.connect(self.fire)
        self.exitButton.clicked.connect(self.ext)
 
    def start_webcam(self):
        global Agecnt
        global exitflag
        global th_flag
        global name
        global xl_row
        Agecnt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        wb = Workbook()
        xl_row = 0
        ws1 = wb.active  # 현재 열려 있는 Sheet
        now = time.gmtime(time.time())
        now_list = list(now)
        name = str(now_list[0]) + "_" + str(now_list[1]) + "_" + str(now_list[2]) + "-" + str(now_list[3]) + "__" + str(
            now_list[4]) + "__" + str(now_list[5])
 
        ws1.title = name
        wb.save("D:/" + ws1.title + ".xlsx")
 
        Agecnt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        exitflag = 0
 
        self.imgtimer.timeout.connect(self.update_frame)
        # self.timer.timeout.connect(self.graph)
        self.imgtimer.start(1000)
        # self.timer.start(1000)
 
        if (th_flag == 0):
            self.th_main = threading.Thread(target=main, args=())
            # self.th_main.daemon = True
            self.th_main.start()
 
        th_flag = 1
        print("start webcam")
 
    def stop_webcam(self):
        global exitflag
        global th_flag
        exitflag = 1
        # self.timer.stop()
        self.imgtimer.stop()
        self.th_main.join()
        if ad_output.timer.is_alive():
            ad_output.timer.cancel()
        if warning_output.timer_w.is_alive():
            warning_output.timer_w.cancel()
        th_flag = 0
        if sock_on == 1:
            sock.send("qqqqqqqq".encode())
            sock.close()
        print("stop webcam")
 
    def update_frame(self):
        height, width, channel = myimge.shape
        bytesPerLine = 3 * width
        qImg = QImage(myimge.data, width, height, bytesPerLine, QImage.Format_RGB888).rgbSwapped()
        pixmap = QPixmap(qImg)
        self.imgLabel.setPixmap(pixmap)
        self.widget.canvas.ax.clear()
        self.widget.canvas.ax.bar(Range, Agecnt)
        self.widget.canvas.draw()
        print("update_frame")
 
    def graph(self):
        self.widget.canvas.ax.clear()
        self.widget.canvas.ax.bar(Range, Agecnt)
        self.widget.canvas.draw()
 
    def reset(self):
        global Agecnt
        global name
        global xl_row
        wb = load_workbook("D:/" + name + ".xlsx")
        ws = wb.active
        xl_row = xl_row + 1
        for col in range(1, 9):
            ws.cell(row=xl_row, column=col, value=int(Agecnt[col - 1]))
 
        wb.save("D:/" + name + ".xlsx")
 
        Agecnt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        print("reset!")
        # sock.send("reset123".encode())
 
    def warning_event(self):
        Agecnt[8] = -1
        warning_output()
        # movie = Movie_MP4(Agevideo[8])
        # movie.play()
 
    def fire(self):
        Agecnt[8] = -2
        # movie = Movie_MP4(Agevideo[8])
        # movie.play()
        warning_output()
 
    def ext(self):
        Agecnt[8] = -3
        warning_output()
 
 
def ad_output():
    global Agecnt
    if Agecnt[8] == 0:
        Max = Agecnt.index(max(Agecnt))
        if not Agecnt[Max] == 0:  # Data 가 있을 경우 해당 나이에 맞는 광고 표출
            path = random.choice(glob.glob(Agevideo[Max] + "\*"))  # 랜덤 파일 추출
            f_name = os.path.split(path)  # 파일명 추출 s[1] -> 파일명
            T = int(f_name[1][0]) * 100 + int(f_name[1][1]) * 10 + int(f_name[1][2])  # 파일명에서 시간 추출
            s_d = os.path.splitext(path)  # mp4 제외 파일명만 추출
            s_d = os.path.split(s_d[0])
 
            if sock_on == 1:
                sock.send(s_d[1].encode())
                print("동영상 제목 전송")
                while 1:
                    receive = sock.recv(8)
                    print(receive)
                    if receive :
                        print("나감")
                        break;
 
            ad_output.timer = threading.Timer(T, ad_output)
            #movie = Movie_MP4(path)
            #movie.play()
            ad_output.timer.start()
            
        else:  # Data 값들이 없을 경우 2초 간격으로 타이머 반복 실행
            ad_output.timer = threading.Timer(2, ad_output)
            ad_output.timer.start()
    else:
        ad_output.timer.cancel()
        print("광고 타이머가 Warning 에 의해 종료되었습니다.")
 
 
def warning_output():
    global Agecnt
    if Agecnt[8] != 0:
        w_index = -1 * Agecnt[8] + 7
        f_name = os.path.split(Agevideo[w_index])  # 파일명 추출 s[1] -> 파일명
        T = int(f_name[1][0]) * 100 + int(f_name[1][1]) * 10 + int(f_name[1][2])  # 파일명에서 시간 추출
 
        s_d = os.path.splitext(Agevideo[w_index])  # mp4 제외 파일명만 추출
        s_d = os.path.split(s_d[0])
        if sock_on == 1:
            sock.send(s_d[1].encode())
            print("Warning 동영상 제목 전송")
 
        warning_output.timer_w = threading.Timer(T, warning_output)
        #movie = Movie_MP4(Agevideo[w_index])
        #movie.play()
        warning_output.timer_w.start()
    else:
        warning_output.timer_w.cancel()
        print("Warning 타이머가 reset에 의해 종료되었습니다.")
 
 
pretrained_model = "https://github.com/yu4u/age-gender-estimation/releases/download/v0.5/weights.18-4.06.hdf5"
modhash = '89f56a39a78454e96379348bddd78c0d'
 
 
def get_args():
    parser = argparse.ArgumentParser(description="This script detects faces from web cam input, "
                                                 "and estimates age and gender for the detected faces.",
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--weight_file", type=str, default=None,
                        help="path to weight file (e.g. weights.18-4.06.hdf5)")
    parser.add_argument("--depth", type=int, default=16,
                        help="depth of network")
    parser.add_argument("--width", type=int, default=8,
                        help="width of network")
    args = parser.parse_args()
    return args
 
 
def draw_label(image, point, label, font=cv2.FONT_HERSHEY_SIMPLEX,
               font_scale=1, thickness=2):
    size = cv2.getTextSize(label, font, font_scale, thickness)[0]
    x, y = point
    cv2.rectangle(image, (x, y - size[1]), (x + size[0], y), (255, 0, 0), cv2.FILLED)
    cv2.putText(image, label, point, font, font_scale, (255, 255, 255), thickness)
 
 
@contextmanager
def video_capture(*args, **kwargs):
    cap = cv2.VideoCapture(*args, **kwargs)
    try:
        yield cap
    finally:
        cap.release()
 
 
def yield_images():
    # capture video
    with video_capture(0) as cap:
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
 
        while True:
            # get video frame
            ret, img = cap.read()
 
            if not ret:
                raise RuntimeError("Failed to capture image")
 
            yield img
 
 
def main():
    global exitflag
    global testflag
    ad_output()
    print("일단 메인문은 들어왔습니다잉")
    args = get_args()
    depth = args.depth
    k = args.width
    weight_file = args.weight_file
    model = 0
    if not weight_file:
        weight_file = get_file("weights.18-4.06.hdf5", pretrained_model, cache_subdir="pretrained_models",
                               file_hash=modhash, cache_dir=os.path.dirname(os.path.abspath(__file__)))
        print("웨이트 파일이 없으므로 받아옵니다.")
        print(weight_file)
    print("웨이트 파일 받습니다.")
    # for face detection
    detector = dlib.get_frontal_face_detector()
    print("얼굴인식 라이브러리 사용")
    # load model and weights
    img_size = 64
    print("이미지 사이즈 = 64")
    model = WideResNet(img_size, depth=depth, k=k)()
    print("wideResNet 설정")
    if testflag == 0:
        model.load_weights(weight_file)
    print("사진 찍기 전!")
    for img in yield_images():
        input_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        img_h, img_w, _ = np.shape(input_img)
 
        # detect faces using dlib detector
        detected = detector(input_img, 1)
        faces = np.empty((len(detected), img_size, img_size, 3))
 
        if len(detected) > 0:
            for i, d in enumerate(detected):
                x1, y1, x2, y2, w, h = d.left(), d.top(), d.right() + 1, d.bottom() + 1, d.width(), d.height()
                xw1 = max(int(x1 - 0.4 * w), 0)
                yw1 = max(int(y1 - 0.4 * h), 0)
                xw2 = min(int(x2 + 0.4 * w), img_w - 1)
                yw2 = min(int(y2 + 0.4 * h), img_h - 1)
                cv2.rectangle(img, (x1, y1), (x2, y2), (255, 0, 0), 2)
                # cv2.rectangle(img, (xw1, yw1), (xw2, yw2), (255, 0, 0), 2)
                faces[i, :, :, :] = cv2.resize(img[yw1:yw2 + 1, xw1:xw2 + 1, :], (img_size, img_size))
 
            # predict ages and genders of the detected faces
            results = model.predict(faces)
            predicted_genders = results[0]
            ages = np.arange(0, 101).reshape(101, 1)
            predicted_ages = results[1].dot(ages).flatten()
 
            # draw results
            for i, d in enumerate(detected):
                label = "{}, {}".format(int(predicted_ages[i]),
                                        "F" if predicted_genders[i][0] > 0.5 else "M")
                draw_label(img, (d.left(), d.top()), label)
                age_index = (int)(predicted_ages[i] / 20)
                if predicted_genders[i][0] > 0.5:
                    if age_index < 3:
                        Agecnt[age_index] = Agecnt[age_index] + 1
                    else:
                        Agecnt[3] = Agecnt[3] + 1
                else:
                    if age_index + 4 < 7:
                        Agecnt[age_index + 4] = Agecnt[age_index + 4] + 1
                    else:
                        Agecnt[7] = Agecnt[7] + 1
 
        global myimge
        myimge = img
 
        print("main 실행")
 
        # cv2.imshow("result", img)
        # key = cv2.waitKey(30)
        # cv2.waitKey(30)
        if exitflag == 1:
            break
        # while(exitflag):
        # time.sleep(2)
    print("main 끝")
    exitflag = 0
    testflag = 1
 
 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Life2Coding()
    window.setWindowTitle('AD title')
 
    window.show()
    app.exec_()
    # sys.exit(app.exec_())
